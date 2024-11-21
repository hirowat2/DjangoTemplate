import csv

# from django.views.generic import ListView
import openpyxl
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_http_methods
from pyexcelerate import Workbook

from .forms import RepositionForm
from .models import Reposition
from .services import csv_to_list_in_memory, save_data

# class RepositionListView(ListView):
#     model = Reposition
#     paginate_by = 5


def reposition_list(request):
    template_name = 'reposition/reposition_list.html'
    object_list = Reposition.objects.all()

    search = request.GET.get('search')

    if search:
        object_list = object_list.filter(
            Q(title__icontains=search)
            | Q(description__icontains=search)
            | Q(category__title__icontains=search)
        )

    # https://docs.djangoproject.com/en/4.1/topics/pagination/#using-paginator-in-a-view-function
    items_per_page = 10
    paginator = Paginator(object_list, items_per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'items_count': page_obj.object_list.count(),
    }
    return render(request, template_name, context)


def reposition_detail(request, pk):
    template_name = 'reposition/reposition_detail.html'
    instance = get_object_or_404(Reposition, pk=pk)

    context = {'object': instance}
    return render(request, template_name, context)


def reposition_create(request):
    template_name = 'reposition/reposition_form.html'
    form = RepositionForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('reposition:reposition_list')

    verbose_name_plural = form.instance._meta.verbose_name_plural
    context = {
        'form': form,
        'verbose_name_plural': verbose_name_plural,
    }
    return render(request, template_name, context)


def reposition_update(request, pk):
    template_name = 'reposition/reposition_form.html'
    instance = get_object_or_404(Reposition, pk=pk)
    form = RepositionForm(request.POST or None, instance=instance)

    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('reposition:reposition_list')

    context = {'form': form, 'object': instance}
    return render(request, template_name, context)


def reposition_delete(request, pk):
    instance = get_object_or_404(Reposition, pk=pk)
    instance.delete()
    return redirect('reposition:reposition_list')


@require_http_methods(['POST'])
def import_view(request):
    filename = request.FILES.get('filename')

    if filename.name.endswith('.csv'):
        data = csv_to_list_in_memory(filename)
    else:
        wb = openpyxl.load_workbook(filename, data_only=True)
        ws = wb.active

        max_row = ws.max_row
        max_col = ws.max_column

        data = []

        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            _dict = dict(title=row[0].value, price=row[1].value)
            data.append(_dict)

    save_data(data)
    return redirect('reposition:reposition_list')


def export_csv(request):
    with open('/tmp/repositions_out.csv', 'w') as f:
        csv_writer = csv.writer(f)
        repositions = reposition.objects.all()

        csv_writer.writerow(['title', 'price'])
        for reposition in repositions:
            csv_writer.writerow([reposition.title, reposition.price])

    return redirect('reposition:reposition_list')


def export_xlsx(request):
    repositions = Reposition.objects.all()
    data = [(reposition.title, reposition.price) for reposition in repositions]
    data.insert(0, ('title', 'price'))

    wb = Workbook()
    wb.new_sheet("sheet name", data=data)
    wb.save("/tmp/repositions_out.xlsx")
    return redirect('reposition:reposition_list')
