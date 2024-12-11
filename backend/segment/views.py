import csv

# from django.views.generic import ListView
import openpyxl
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_http_methods
from pyexcelerate import Workbook

from .forms import SegmentForm
from .models import Segment
from .services import csv_to_list_in_memory, save_data

# class SegmentListView(ListView):
#     model = Segment
#     paginate_by = 5


def segment_list(request):
    template_name = 'segment/segment_list.html'
    object_list = Segment.objects.all()

    search = request.GET.get('search')

    if search:
        object_list = object_list.filter(
            Q(product__title__icontains=search)
            | Q(consumo_historico__icontains=search)
            | Q(cv_diario__icontains=search)
        )


    # https://docs.djangoproject.com/en/4.1/topics/pagination/#using-paginator-in-a-view-function
    items_per_page = 10
    paginator = Paginator(object_list, items_per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'items_count': page_obj.object_list.count(),
    }
    return render(request, template_name, context)


def segment_detail(request, pk):
    template_name = 'segment/segment_detail.html'
    instance = get_object_or_404(Segment, pk=pk)

    context = {'object': instance}
    return render(request, template_name, context)


def segment_create(request):
    template_name = 'segment/segment_form.html'
    form = SegmentForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        segment = form.save(commit=False)
        # Defina o produto se ele não estiver no formulário
        # segment.product = Product.objects.first()  # ou outro valor válido
        segment.save()
        return redirect('segment:segment_list')

    verbose_name_plural = form.instance._meta.verbose_name_plural
    context = {
        'form': form,
        'verbose_name_plural': verbose_name_plural,
    }
    return render(request, template_name, context)


def segment_update(request, pk):
    template_name = 'segment/segment_form.html'
    instance = get_object_or_404(Segment, pk=pk)
    form = SegmentForm(request.POST or None, instance=instance)

    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('segment:segment_list')

    context = {'form': form, 'object': instance}
    return render(request, template_name, context)


def segment_delete(request, pk):
    instance = get_object_or_404(Segment, pk=pk)
    instance.delete()
    return redirect('segment:segment_list')


@require_http_methods(['POST'])
def import_view(request):
    filename = request.FILES.get('filename')

    if filename.name.endswith('.csv'):
        data = csv_to_list_in_memory(filename)
    else:
        wb = openpyxl.load_workbook(filename, data_only=True)
        ws = wb.active

        max_row = ws.max_row
        max_col = ws.max_column

        data = []

        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            _dict = dict(title=row[0].value, price=row[1].value, product=Product.objects.get(title=title))
            data.append(_dict)

    save_data(data)
    return redirect('segment:segment_list')


def export_csv(request):
    with open('/tmp/segment_out.csv', 'w') as f:
        csv_writer = csv.writer(f)
        segments = segment.objects.all()

        csv_writer.writerow(['title', 'price'])
        for segment in segments:
            csv_writer.writerow([segment.title])

    return redirect('segment:segment_list')


def export_xlsx(request):
    segments = Segment.objects.all()
    data = [(segment.title) for segment in segments]
    data.insert(0, ('title', 'price'))

    wb = Workbook()
    wb.new_sheet("sheet name", data=data)
    wb.save("/tmp/segment_out.xlsx")
    return redirect('segment:segment_list')