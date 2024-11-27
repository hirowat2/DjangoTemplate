import csv

# from django.views.generic import ListView
import openpyxl
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_http_methods
from pyexcelerate import Workbook

from .forms import PrateleiraForm
from .models import Prateleira
from .services import csv_to_list_in_memory, save_data

from backend.product.models import Product  # Ajuste o caminho para o modelo Product

# class prateleiraListView(ListView):
#     model = prateleira
#     paginate_by = 5


def prateleira_list(request):
    template_name = 'prateleira/prateleira_list.html'
    object_list = Prateleira.objects.all()

    search = request.GET.get('search')

    if search:
        object_list = object_list.filter(
            Q(product__title__icontains=search)
            | Q(consumo_historico__icontains=search)
            | Q(cv_diario__icontains=search)
        )


    # https://docs.djangoproject.com/en/4.1/topics/pagination/#using-paginator-in-a-view-function
    items_per_page = 10
    paginator = Paginator(object_list, items_per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'items_count': page_obj.object_list.count(),
    }
    return render(request, template_name, context)


def prateleira_detail(request, pk):
    template_name = 'prateleira/prateleira_detail.html'
    instance = get_object_or_404(Prateleira, pk=pk)

    context = {'object': instance}
    return render(request, template_name, context)


def prateleira_create(request):
    template_name = 'prateleira/prateleira_form.html'
    form = PrateleiraForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        prateleira = form.save(commit=False)
        # Defina o produto se ele não estiver no formulário
        # prateleira.product = Product.objects.first()  # ou outro valor válido
        prateleira.save()
        return redirect('prateleira:prateleira_list')

    verbose_name_plural = form.instance._meta.verbose_name_plural
    context = {
        'form': form,
        'verbose_name_plural': verbose_name_plural,
    }
    return render(request, template_name, context)


def prateleira_update(request, pk):
    template_name = 'prateleira/prateleira_form.html'
    instance = get_object_or_404(Prateleira, pk=pk)
    form = PrateleiraForm(request.POST or None, instance=instance)

    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('prateleira:prateleira_list')

    context = {'form': form, 'object': instance}
    return render(request, template_name, context)


def prateleira_delete(request, pk):
    instance = get_object_or_404(Prateleira, pk=pk)
    instance.delete()
    return redirect('prateleira:prateleira_list')


@require_http_methods(['POST'])
def import_view(request):
    filename = request.FILES.get('filename')

    if filename.name.endswith('.csv'):
        data = csv_to_list_in_memory(filename)
    else:
        wb = openpyxl.load_workbook(filename, data_only=True)
        ws = wb.active

        max_row = ws.max_row
        max_col = ws.max_column

        data = []

        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            _dict = dict(title=row[0].value, price=row[1].value, product=Product.objects.get(title=title))
            data.append(_dict)

    save_data(data)
    return redirect('prateleira:prateleira_list')


def export_csv(request):
    with open('/tmp/prateleiras_out.csv', 'w') as f:
        csv_writer = csv.writer(f)
        prateleiras = prateleira.objects.all()

        csv_writer.writerow(['title', 'price'])
        for prateleira in prateleiras:
            csv_writer.writerow([prateleira.consumo_historico, prateleira.demanda_dia_prev])

    return redirect('prateleira:prateleira_list')


def export_xlsx(request):
    prateleiras = Prateleira.objects.all()
    data = [(prateleira.consumo_historico, prateleira.demanda_dia_prev) for prateleira in prateleiras]
    data.insert(0, ('title', 'price'))

    wb = Workbook()
    wb.new_sheet("sheet name", data=data)
    wb.save("/tmp/prateleiras_out.xlsx")
    return redirect('prateleira:prateleira_list')
